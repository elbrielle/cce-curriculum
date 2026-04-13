#!/usr/bin/env python3
"""
build_xlsx.py -- Rebuild the CCE Scope & Sequence spreadsheet from markdown sources.

Usage:
    python build/build_xlsx.py

Produces output/CCE_Comprehensive_Scope_Sequence.xlsx with 4 tabs:
  1. Master Pacing Guide      (from scope-and-sequence.md)
  2. TEKS Coverage Matrix     (from resources/teks-coverage-matrix.md)
  3. Free Resource Directory  (from resources/free-resource-directory.md)
  4. eDynamic Unit Map        (from resources/edynamic-unit-map.md)
"""

import sys, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build"))
from config import XLSX_TAB_COLORS

RESOURCES = ROOT / "cce-curriculum" / "resources"
OUTPUT_DIR = ROOT / "output"
OUTPUT_FILE = OUTPUT_DIR / "CCE_Comprehensive_Scope_Sequence.xlsx"

# ── Styles ──
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
BODY_FONT = Font(name="Arial", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Alternating row fills
ROW_FILL_A = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ROW_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def parse_md_table(filepath):
    """Parse a markdown file containing a pipe-delimited table.
    Returns (headers, rows) where each row is a list of cell strings.
    """
    text = Path(filepath).read_text(encoding="utf-8")
    lines = text.strip().split('\n')

    # Find table lines (start with |)
    table_lines = [l for l in lines if l.strip().startswith('|')]
    if len(table_lines) < 2:
        raise ValueError(f"No table found in {filepath}")

    # First line = headers, second = separator, rest = data
    def split_row(line):
        cells = line.strip().strip('|').split('|')
        return [c.strip() for c in cells]

    headers = split_row(table_lines[0])

    rows = []
    for line in table_lines[2:]:  # skip header + separator
        row = split_row(line)
        # Pad or trim to match header count
        while len(row) < len(headers):
            row.append("")
        rows.append(row[:len(headers)])

    return headers, rows


def add_sheet(wb, name, headers, rows, col_widths=None, tab_color=None):
    """Add a formatted worksheet to the workbook."""
    ws = wb.create_sheet(title=name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write headers
    for ci, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for ri, row in enumerate(rows, 2):
        fill = ROW_FILL_A if ri % 2 == 0 else ROW_FILL_B
        for ci, value in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.fill = fill
            cell.border = THIN_BORDER

    # Set column widths
    if col_widths:
        for ci, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = width
    else:
        # Auto-size: min 12, max 50, based on header length + padding
        for ci, header in enumerate(headers, 1):
            w = min(max(len(header) + 4, 14), 50)
            ws.column_dimensions[get_column_letter(ci)].width = w

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    return ws


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Tab 1: Master Pacing Guide ──
    src = ROOT / "cce-curriculum" / "scope-and-sequence.md"
    headers, rows = parse_md_table(src)
    add_sheet(wb, "Master Pacing Guide", headers, rows,
              col_widths=[14, 8, 22, 20, 40, 45, 18, 16, 20, 35, 14, 40, 30],
              tab_color=XLSX_TAB_COLORS["Master Pacing Guide"])

    # ── Tab 2: TEKS Coverage Matrix ──
    src = RESOURCES / "teks-coverage-matrix.md"
    headers, rows = parse_md_table(src)
    add_sheet(wb, "TEKS Coverage Matrix", headers, rows,
              col_widths=[12, 45, 35, 30, 16],
              tab_color=XLSX_TAB_COLORS["TEKS Coverage Matrix"])

    # ── Tab 3: Free Resource Directory ──
    src = RESOURCES / "free-resource-directory.md"
    headers, rows = parse_md_table(src)
    add_sheet(wb, "Free Resource Directory", headers, rows,
              col_widths=[28, 55, 16, 40, 10, 25],
              tab_color=XLSX_TAB_COLORS["Free Resource Directory"])

    # ── Tab 4: eDynamic Unit Map ──
    src = RESOURCES / "edynamic-unit-map.md"
    headers, rows = parse_md_table(src)
    add_sheet(wb, "eDynamic Unit Map", headers, rows,
              col_widths=[14, 40, 25, 45, 22, 18],
              tab_color=XLSX_TAB_COLORS["eDynamic Unit Map"])

    # ── Save ──
    wb.save(str(OUTPUT_FILE))
    print(f"Built {OUTPUT_FILE}")
    print(f"  Tabs: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
